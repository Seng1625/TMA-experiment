from __future__ import annotations

import os
import re
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List

from flask import Flask, jsonify, render_template, request, send_file, Response

app = Flask(__name__)

def _safe_filename(name: str) -> str:
    name = re.sub(r"[^a-zA-Z0-9._-]+", "_", name).strip("._-")
    return name or "results"

def _build_workbook(participant_id: str, session_started: str, results: List[Dict[str, Any]]):
    """Build a readable Excel workbook with colors + feedback phases."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

    wb = Workbook()
    ws = wb.active
    ws.title = "responses"

    headers = [
        "participantId",
        "sessionStartedAt",
        "questionnaire",          # Q1 / Q2
        "blockSize",
        "blockIndex",
        "trialInBlock",
        "feedbackPhase",          # Q2 only (blockIndex-1)
        "question",
        "type",
        "title",
        "prompt",
        "options",
        "correctAnswer",
        "aiSuggestion",
        "aiConfidence",
        "aiCorrect",
        "initialChoice",
        "initialConfidence",
        "initialCorrect",
        "finalChoice",
        "finalConfidence",
        "finalCorrect",
        "changedAfterAI",
        "followedAI_final",
        "initialRTms",
        "finalRTms",
        "timestamp",
    ]
    ws.append(headers)

    # write rows
    for r in results:
        mode = str(r.get("mode", ""))
        block_index = int(r.get("blockIndex", 1) or 1)
        trial_in_block = int(r.get("trialInBlock", 1) or 1)
        feedback_phase = ""
        if mode == "Q2":
            feedback_phase = max(block_index - 1, 0)

        ws.append([
            participant_id,
            session_started,
            mode,
            r.get("blockSize", ""),
            block_index,
            trial_in_block,
            feedback_phase,
            r.get("trialIndex", ""),
            r.get("type", ""),
            r.get("title", ""),
            r.get("prompt", ""),
            ", ".join(r.get("options", []) or []),
            r.get("correct", ""),
            r.get("aiSuggestion", ""),
            r.get("aiConfidence", ""),
            bool(r.get("aiIsCorrect", False)),
            r.get("initialChoice", ""),
            r.get("initialConfidence", ""),
            bool(r.get("initialCorrect", False)),
            r.get("finalChoice", ""),
            r.get("finalConfidence", ""),
            bool(r.get("finalCorrect", False)),
            bool(r.get("changed", False)),
            bool(r.get("followedAI_final", False)),
            r.get("initialRTms", ""),
            r.get("finalRTms", ""),
            r.get("timestamp", ""),
        ])

    # ---------- Formatting ----------
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="2B2F3A")
    thick = Side(style="medium", color="9CA3AF")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = thin_border

    # body borders + wrap prompt
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            if col == headers.index("prompt") + 1:
                c.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                c.alignment = Alignment(vertical="top")

    header_to_col = {h: i for i, h in enumerate(headers, start=1)}
    def col_letter(h: str) -> str:
        return get_column_letter(header_to_col[h])

    # Block separators (thick top border when trialInBlock==1)
    tib_col = header_to_col["trialInBlock"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=tib_col).value == 1:
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=row, column=col)
                c.border = Border(left=thin, right=thin, top=thick, bottom=thin)

    # Colors
    green = PatternFill("solid", fgColor="D1FAE5")
    red = PatternFill("solid", fgColor="FEE2E2")
    yellow = PatternFill("solid", fgColor="FEF3C7")
    blue = PatternFill("solid", fgColor="DBEAFE")

    # Correct columns
    for h in ("aiCorrect", "initialCorrect", "finalCorrect"):
        rng = f"{col_letter(h)}2:{col_letter(h)}{ws.max_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["TRUE"], fill=green))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["FALSE"], fill=red))

    # Changed / Followed
    rng_changed = f"{col_letter('changedAfterAI')}2:{col_letter('changedAfterAI')}{ws.max_row}"
    ws.conditional_formatting.add(rng_changed, CellIsRule(operator="equal", formula=["TRUE"], fill=yellow))

    rng_follow = f"{col_letter('followedAI_final')}2:{col_letter('followedAI_final')}{ws.max_row}"
    ws.conditional_formatting.add(rng_follow, CellIsRule(operator="equal", formula=["TRUE"], fill=blue))

    # Confidence columns: heatmap
    for h in ("aiConfidence", "initialConfidence", "finalConfidence"):
        rng = f"{col_letter(h)}2:{col_letter(h)}{ws.max_row}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="num", start_value=50, start_color="FEE2E2",
            mid_type="num", mid_value=75, mid_color="FEF3C7",
            end_type="num", end_value=100, end_color="D1FAE5"
        ))

    # feedbackPhase colored (helps to see before/after feedback)
    phase_col = header_to_col["feedbackPhase"]
    palette = ["E0F2FE", "EDE9FE", "DCFCE7", "FFF7ED", "FCE7F3", "ECFCCB"]
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=phase_col).value
        if val is None or val == "":
            continue
        try:
            v = int(val)
        except Exception:
            continue
        ws.cell(row=row, column=phase_col).fill = PatternFill("solid", fgColor=palette[v % len(palette)])

    # Column widths (some fixed + auto for the rest)
    preset = {
        "participantId": 14,
        "sessionStartedAt": 22,
        "questionnaire": 14,
        "blockSize": 10,
        "blockIndex": 10,
        "trialInBlock": 11,
        "feedbackPhase": 13,
        "question": 9,
        "type": 18,
        "title": 22,
        "prompt": 46,
        "options": 22,
        "correctAnswer": 14,
        "aiSuggestion": 14,
        "aiConfidence": 12,
        "initialChoice": 14,
        "initialConfidence": 14,
        "finalChoice": 12,
        "finalConfidence": 13,
        "changedAfterAI": 14,
        "followedAI_final": 15,
        "initialRTms": 11,
        "finalRTms": 10,
        "timestamp": 22,
    }
    for h, w in preset.items():
        ws.column_dimensions[col_letter(h)].width = w
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].width:
            continue
        max_len = 10
        for row in range(1, min(ws.max_row, 300) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)[:80]))
        ws.column_dimensions[letter].width = min(60, max_len + 2)

    # ---------- Summary sheets ----------
    ws2 = wb.create_sheet("summary")
    ws2.append(["participantId", participant_id])
    ws2.append(["sessionStartedAt", session_started])
    ws2.append(["questions", len(results)])

    if results:
        changed = sum(1 for r in results if r.get("changed"))
        follow_final = sum(1 for r in results if r.get("followedAI_final"))
        init_acc = sum(1 for r in results if r.get("initialCorrect"))
        final_acc = sum(1 for r in results if r.get("finalCorrect"))
        ai_acc = sum(1 for r in results if r.get("aiIsCorrect"))

        ws2.append(["%changedAfterAI", round(100 * changed / len(results), 1)])
        ws2.append(["%followAI_final", round(100 * follow_final / len(results), 1)])
        ws2.append(["%accuracy_initial", round(100 * init_acc / len(results), 1)])
        ws2.append(["%accuracy_final", round(100 * final_acc / len(results), 1)])
        ws2.append(["%accuracy_AI", round(100 * ai_acc / len(results), 1)])

    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.freeze_panes = "A2"

    # blocks sheet for before/after feedback tracking
    ws3 = wb.create_sheet("blocks")
    ws3.append(["questionnaire", "blockSize", "blockIndex", "feedbackPhase", "startQuestion", "endQuestion",
                "n", "%acc_initial", "%acc_final", "%acc_AI", "%changed", "%followAI_final"])

    blocks = {}
    mode_any = ""
    block_size_any = ""
    for r in results:
        mode_any = r.get("mode", mode_any)
        block_size_any = r.get("blockSize", block_size_any)
        b = int(r.get("blockIndex", 1) or 1)
        blocks.setdefault(b, []).append(r)

    for b in sorted(blocks.keys()):
        rr = blocks[b]
        n = len(rr)
        start_q = min(int(x.get("trialIndex", 0) or 0) for x in rr)
        end_q = max(int(x.get("trialIndex", 0) or 0) for x in rr)
        fb_phase = max(b-1,0) if mode_any == "Q2" else ""
        ws3.append([
            mode_any,
            block_size_any,
            b,
            fb_phase,
            start_q,
            end_q,
            n,
            round(100*sum(1 for x in rr if x.get("initialCorrect"))/n, 1),
            round(100*sum(1 for x in rr if x.get("finalCorrect"))/n, 1),
            round(100*sum(1 for x in rr if x.get("aiIsCorrect"))/n, 1),
            round(100*sum(1 for x in rr if x.get("changed"))/n, 1),
            round(100*sum(1 for x in rr if x.get("followedAI_final"))/n, 1),
        ])

    for cell in ws3[1]:
        cell.font = Font(bold=True)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(ws3.max_column)}{ws3.max_row}"

    # dictionary sheet
    ws4 = wb.create_sheet("dictionary")
    ws4.append(["column", "meaning"])
    explanations = {
        "questionnaire": "Q1 = sans feedback, Q2 = avec feedback toutes les 8 questions (par défaut).",
        "blockIndex": "Numéro de bloc (ex: 1..5 si 40 questions et bloc=8).",
        "trialInBlock": "Position de la question à l’intérieur du bloc (1..blockSize).",
        "feedbackPhase": "Q2 uniquement : nombre d’écrans de feedback déjà vus avant cette question (blocIndex-1).",
        "aiCorrect": "TRUE si la réponse de l’IA est correcte.",
        "initialCorrect": "TRUE si la réponse AVANT IA est correcte.",
        "finalCorrect": "TRUE si la réponse APRÈS IA est correcte.",
        "changedAfterAI": "TRUE si le participant a changé sa réponse après avoir vu l’IA.",
        "followedAI_final": "TRUE si la réponse finale = suggestion de l’IA.",
        "initialRTms": "Temps de réponse avant IA (ms).",
        "finalRTms": "Temps de réponse après IA (ms).",
        "aiConfidence": "Confiance de l’IA (en %).",
        "initialConfidence": "Confiance du participant avant IA (en %).",
        "finalConfidence": "Confiance du participant après IA (en %).",
    }
    for h in headers:
        ws4.append([h, explanations.get(h, "")])

    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 95
    ws4.freeze_panes = "A2"
    for cell in ws4[1]:
        cell.font = Font(bold=True)

    return wb
def _build_csv(participant_id: str, session_started: str, results: List[Dict[str, Any]]) -> str:
    import csv, io
    headers = [
        "participantId","sessionStartedAt","questionnaire","blockSize","blockIndex","trialInBlock","feedbackPhase",
        "question","type","title","prompt","options","correctAnswer",
        "aiSuggestion","aiConfidence","aiCorrect",
        "initialChoice","initialConfidence","initialCorrect",
        "finalChoice","finalConfidence","finalCorrect",
        "changedAfterAI","followedAI_final","initialRTms","finalRTms","timestamp"
    ]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in results:
        mode = str(r.get("mode", ""))
        b = int(r.get("blockIndex", 1) or 1)
        feedback_phase = "" if mode != "Q2" else max(b - 1, 0)
        w.writerow([
            participant_id, session_started,
            mode,
            r.get("blockSize",""),
            b,
            r.get("trialInBlock",""),
            feedback_phase,
            r.get("trialIndex",""),
            r.get("type",""),
            r.get("title",""),
            r.get("prompt",""),
            ", ".join(r.get("options",[]) or []),
            r.get("correct",""),
            r.get("aiSuggestion",""),
            r.get("aiConfidence",""),
            bool(r.get("aiIsCorrect",False)),
            r.get("initialChoice",""),
            r.get("initialConfidence",""),
            bool(r.get("initialCorrect",False)),
            r.get("finalChoice",""),
            r.get("finalConfidence",""),
            bool(r.get("finalCorrect",False)),
            bool(r.get("changed",False)),
            bool(r.get("followedAI_final",False)),
            r.get("initialRTms",""),
            r.get("finalRTms",""),
            r.get("timestamp",""),
        ])
    return buf.getvalue()
@app.get("/")
def home():
    return render_template("prototype.html", preset_mode="")

@app.get("/q1")
def q1():
    return render_template("prototype.html", preset_mode="Q1")

@app.get("/q2")
def q2():
    return render_template("prototype.html", preset_mode="Q2")

@app.get("/types")
def types():
    return render_template("prototype.html", preset_mode="TYPES")

@app.post("/api/export")
def export_xlsx():
    try:
        payload: Dict[str, Any] = request.get_json(force=True, silent=False)  # type: ignore
        participant_id = str(payload.get("participantId", "P00000"))
        session_started = str(payload.get("sessionStartedAt", datetime.utcnow().isoformat()))
        results: List[Dict[str, Any]] = payload.get("results", []) or []

        wb = _build_workbook(participant_id, session_started, results)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = _safe_filename(f"results_{participant_id}_{ts}.xlsx")

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            max_age=0,
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.post("/api/export_csv")
def export_csv():
    try:
        payload: Dict[str, Any] = request.get_json(force=True, silent=False)  # type: ignore
        participant_id = str(payload.get("participantId", "P00000"))
        session_started = str(payload.get("sessionStartedAt", datetime.utcnow().isoformat()))
        results: List[Dict[str, Any]] = payload.get("results", []) or []
        csv_text = _build_csv(participant_id, session_started, results)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = _safe_filename(f"results_{participant_id}_{ts}.csv")
        return Response(
            csv_text,
            mimetype="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8000"))
    app.run(host="0.0.0.0", port=port, debug=False)
